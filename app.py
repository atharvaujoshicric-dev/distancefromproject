import streamlit as st
import pandas as pd
import re
import io
import time
from openpyxl.styles import Alignment, PatternFill, Border, Side

# Selenium Imports
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# --- REFINED RERA SCRAPING LOGIC ---
def fetch_rera_details(project_name):
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    # Spoofer: Makes the server think a real human is visiting
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
    
    driver = None
    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        wait = WebDriverWait(driver, 20)
        
        # 1. Navigate to MahaRERA
        driver.get("https://maharerait.mahaonline.gov.in/SearchList/Search")
        
        # 2. Select 'Registered Projects'
        radio = wait.until(EC.element_to_be_clickable((By.ID, "Promoter")))
        driver.execute_script("arguments[0].click();", radio)
        
        # 3. Search Project
        input_box = driver.find_element(By.ID, "ProjectName")
        input_box.send_keys(project_name)
        driver.find_element(By.ID, "btnSearch").click()
        time.sleep(4)
        
        # 4. Click 'View Details' (Using JS click to avoid overlays)
        view_links = driver.find_elements(By.LINK_TEXT, "View Details")
        if not view_links:
            driver.quit()
            return ["N/A"] * 5
        driver.execute_script("arguments[0].click();", view_links[0])
        time.sleep(5)
        
        # 5. WINDOW SWITCHING (Crucial for RERA)
        main_window = driver.current_window_handle
        for handle in driver.window_handles:
            if handle != main_window:
                driver.switch_to.window(handle)
                break
        
        # 6. Extraction using highly flexible XPATHs
        def get_text(xpath):
            try: return wait.until(EC.presence_of_element_located((By.XPATH, xpath))).text
            except: return "N/A"

        possession = get_text("//td[contains(text(),'Proposed Date of Completion')]/following-sibling::td")
        
        # Count Buildings
        towers_list = driver.find_elements(By.XPATH, "//table[@id='BuildingDetails']//tr")
        towers = str(len(towers_list) - 1) if len(towers_list) > 1 else "N/A"
        
        # Calculate Units and Floors from the table columns
        units_total = 0
        max_flr = 0
        unit_cells = driver.find_elements(By.XPATH, "//table[@id='BuildingDetails']//td[4]")
        floor_cells = driver.find_elements(By.XPATH, "//table[@id='BuildingDetails']//td[5]")
        
        for u in unit_cells:
            if u.text.isdigit(): units_total += int(u.text)
        for f in floor_cells:
            if f.text.isdigit(): max_flr = max(max_flr, int(f.text))
            
        units = str(units_total) if units_total > 0 else "N/A"
        floors = f"G+{max_flr}" if max_flr > 0 else "N/A"
        
        # Count Amenities from Development Table
        amn_rows = driver.find_elements(By.XPATH, "//div[@id='DivDevelopmentWork']//tr")
        amenities = f"{len(amn_rows)}+" if len(amn_rows) > 1 else "N/A"

        driver.quit()
        return [amenities, towers, floors, units, possession]
    
    except Exception:
        if driver: driver.quit()
        return ["Not Found"] * 5

# --- REMAINING AREA EXTRACTION LOGIC (Keep as is) ---
def extract_area_logic(text):
    if pd.isna(text) or text == "": return 0.0
    text = " ".join(str(text).split()).replace(' ,', ',').replace(', ', ',')
    m_unit = r'(?:चौ\.?\s*मी\.?|चौरस\s*मी[टत]र|sq\.?\s*m(?:tr)?\.?)'
    f_unit = r'(?:चौ\.?\s*फू\.?|चौरस\s*फु[टत]|sq\.?\s*f(?:t)?\.?)'
    total_keywords = r'(?:ए[ककु]ण\s*क्षेत्र|क्षेत्रफळ|total\s*area)'
    m_segments = re.split(f'(\d+\.?\d*)\s*{m_unit}', text, flags=re.IGNORECASE)
    m_vals = []
    for i in range(1, len(m_segments), 2):
        val, context = float(m_segments[i]), m_segments[i-1].lower()
        if 0 < val < 500 and not any(w in context for w in ["पार्किंग", "पार्कींग", "parking"]):
            m_vals.append(val)
    return round(sum(m_vals), 3) if m_vals else 0.0

def determine_config(area, t1, t2, t3):
    if area == 0: return "N/A"
    return "1 BHK" if area < t1 else "2 BHK" if area < t2 else "3 BHK" if area < t3 else "4 BHK"

def apply_excel_formatting(df, writer, sheet_name, is_summary=True, show_extra=False):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]
    center = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    colors = ["A2D2FF", "FFD6A5", "CAFFBF", "FDFFB6", "FFADAD", "BDB2FF", "9BF6FF"]

    for r in range(1, worksheet.max_row + 1):
        for c in range(1, worksheet.max_column + 1):
            worksheet.cell(row=r, column=c).alignment = center
            if is_summary: worksheet.cell(row=r, column=c).border = border

    if is_summary:
        color_idx, start_prop, start_cfg = 0, 2, 2
        for i in range(2, len(df) + 2):
            curr_p, next_p = df.iloc[i-2, 0], df.iloc[i-1, 0] if i-1 < len(df) else None
            fill = PatternFill(start_color=colors[color_idx % len(colors)], end_color=colors[color_idx % len(colors)], fill_type="solid")
            for col in range(1, len(df.columns) + 1): worksheet.cell(row=i, column=col).fill = fill
            
            if curr_p != next_p:
                m_cols = [1]
                if show_extra: m_cols.extend(range(len(df.columns)-4, len(df.columns)+1))
                for c_idx in m_cols:
                    if i > start_prop: worksheet.merge_cells(start_row=start_prop, start_column=c_idx, end_row=i, end_column=c_idx)
                color_idx, start_prop = color_idx + 1, i + 1
            
            if [df.iloc[i-2, 0], df.iloc[i-2, 1]] != ([df.iloc[i-1, 0], df.iloc[i-1, 1]] if i-1 < len(df) else None):
                if i > start_cfg: worksheet.merge_cells(start_row=start_cfg, start_column=2, end_row=i, end_column=2)
                start_cfg = i + 1

# --- STREAMLIT APP ---
st.set_page_config(page_title="RERA Analysis Tool", layout="wide")
st.title("🏠 Real Estate Data Extractor & Scraper")

st.sidebar.header("Configuration")
loading = st.sidebar.number_input("Loading Factor", value=1.35)
t1 = st.sidebar.number_input("1 BHK <", value=600); t2 = st.sidebar.number_input("2 BHK <", value=850); t3 = st.sidebar.number_input("3 BHK <", value=1100)
show_extra = st.sidebar.checkbox("Fetch Live RERA Data")

uploaded_file = st.file_uploader("Upload Raw Excel", type="xlsx")

if uploaded_file:
    df = pd.read_excel(uploaded_file)
    clean_cols = {c.lower().strip(): c for c in df.columns}
    desc, cons, prop = clean_cols.get('property description'), clean_cols.get('consideration value'), clean_cols.get('property')
    
    if desc and cons and prop:
        with st.spinner('Calculating Area...'):
            df['Carpet Area (SQ.MT)'] = df[desc].apply(extract_area_logic)
            df['Carpet Area (SQ.FT)'] = (df['Carpet Area (SQ.MT)'] * 10.764).round(3)
            df['Saleable Area'] = (df['Carpet Area (SQ.FT)'] * loading).round(3)
            df['APR'] = df.apply(lambda r: round(r[cons]/r['Saleable Area'], 3) if r['Saleable Area'] > 0 else 0, axis=1)
            df['Configuration'] = df['Carpet Area (SQ.FT)'].apply(lambda x: determine_config(x, t1, t2, t3))
        
        valid_df = df[df['Carpet Area (SQ.FT)'] > 0].sort_values([prop, 'Configuration', 'Carpet Area (SQ.FT)'])
        summary = valid_df.groupby([prop, 'Configuration', 'Carpet Area (SQ.FT)']).agg(
            Min_APR=('APR', 'min'), Max_APR=('APR', 'max'), Avg_APR=('APR', 'mean'),
            Median_APR=('APR', 'median'), Mode_APR=('APR', lambda x: x.mode().iloc[0] if not x.mode().empty else 0),
            Property_Count=(prop, 'count')
        ).reset_index()
        summary.columns = ['Property', 'Configuration', 'Carpet Area(SQ.FT)', 'Min. APR', 'Max APR', 'Average of APR', 'Median of APR', 'Mode of APR', 'Count of Property']

        if show_extra:
            unique_projects = summary['Property'].unique()
            project_map = {}
            for p in unique_projects:
                with st.spinner(f"Scraping RERA for: {p}"):
                    project_map[p] = fetch_rera_details(p)
            
            extra_cols = ["Amenities", "Towers", "Floors", "Total Units", "Possession"]
            for i, col in enumerate(extra_cols):
                summary[col] = summary['Property'].apply(lambda x: project_map[x][i])

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Raw Data', index=False)
            raw_ws = writer.sheets['Raw Data']
            for r in range(1, raw_ws.max_row+1):
                for c in range(1, raw_ws.max_column+1):
                    raw_ws.cell(row=r, column=c).alignment = Alignment(horizontal='center', vertical='center')
            
            apply_excel_formatting(summary, writer, 'Summary', show_extra=show_extra)
        
        st.success("Analysis Complete!")
        st.dataframe(summary)
        st.download_button("📥 Download Excel", data=output.getvalue(), file_name="Property_Report.xlsx")
